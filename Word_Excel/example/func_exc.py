from openpyxl import Workbook, load_workbook
# создаём пустой файл
wb = Workbook()
dest_filename = 'empty_book.xlsx'
# создаём в документе новый лист с названием Test
ws2 = wb.create_sheet(title="Test")
# в ячейку F5 записываем значение:
ws2['F5'] = 3.14
# другой вариант записи в ячейку:
ws2.cell(column=5, row=4, value="12345")
# сохраняем новый документ в файле:
wb.save(filename=dest_filename)
# загружаем существующий файл:
wb = load_workbook(filename=dest_filename)
# выбираем нужный лист:
ws = wb['Test']
# получаем значение нужной ячейки:
print(ws['F5'].value)
