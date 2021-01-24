import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment
list_doc = Workbook()
exc_filename = 'matrix.xlsx'
list_excel = list_doc.create_sheet(title='Matrix')
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for x in range(1, 13):
    for y in range(1, 13):
        list_excel.cell(column=x, row=y, value=random.randint(-1000, 1000))
        list_excel.cell(column=x, row=y).border = thin_border
        list_excel.cell(column=x, row=y).alignment = Alignment(horizontal='center')

list_doc.save(exc_filename)

